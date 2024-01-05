import concurrent.futures
from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QGraphicsView, QLineEdit, QPushButton, QLabel,QGraphicsScene, QMessageBox, QProgressDialog,QBoxLayout
import timeit
from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtCore import QFile,Qt , QThread, pyqtSignal,QPropertyAnimation
from PyQt6.QtGui import QPixmap, QImageReader, QMovie, QCursor
import sys,time
import numpy as np
import math
from blessed import Terminal
import openpyxl
import pandas as pd
import os

class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1091, 640)
        self.centralwidget = QtWidgets.QWidget(parent=MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        
        
        palette = QtGui.QPalette()
        palette.setColor(QtGui.QPalette.ColorRole.Window, QtGui.QColor(0, 0, 0))  
        palette.setColor(QtGui.QPalette.ColorRole.WindowText, QtGui.QColor(255, 255, 255))  
        MainWindow.setPalette(palette)
        
        self.p_matris = None
        self.a_matris = None
        self.l_matris = None
        
    
        self.ATP_matris = None
        self.N_matris = None
        self.n_matris = None
        self.m0_matris = None
        self.mX_matris = None
        self.mY_matris = None
        self.mZ_matris = None
        self.Neksi1_matris = None
        self.x_matris = None
        self.v_matris = None
        self.vTPv_matris = None
        self.ATPv_matris = None
        
    
        
        #HESAP KONTROLÜ
        self.p_matriskontrol = 0
        self.a_matriskontrol = 0
        self.l_matriskontrol = 0
        self.goruntule = None
        
        
        
        
        #SOLDAKİ TABLO
        self.tableWidget = QtWidgets.QTableWidget(parent=self.centralwidget)
        self.tableWidget.setGeometry(QtCore.QRect(0, 0, 581, 341))
        self.tableWidget.setObjectName("tableWidget")

        self.tableWidget.setColumnCount(0)
        self.tableWidget.setRowCount(0)
        # Sütun ve satır sayısını başlat
        self.columnCount = 3
        self.rowCount = 3
      
      
      
        #SÜTUN SİL
        self.pushButton_6 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_6.setGeometry(QtCore.QRect(540, 340, 41, 31))
        self.pushButton_6.setObjectName("pushButton_6")
        self.pushButton_6.clicked.connect(self.removeColumn)
        
        #SÜTUN EKLE
        self.pushButton_4 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_4.setGeometry(QtCore.QRect(500, 340, 41, 31))
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton_4.clicked.connect(self.addColumn)
        
        #SATIR EKLE
        self.pushButton_5 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_5.setGeometry(QtCore.QRect(580, 0, 31, 31))
        self.pushButton_5.setObjectName("pushButton_5")
        self.pushButton_5.clicked.connect(self.addRow)
        
        #SATIR SİL
        self.pushButton_7 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_7.setGeometry(QtCore.QRect(580, 30, 31, 31))
        self.pushButton_7.setObjectName("pushButton_7")
        self.pushButton_7.clicked.connect(self.removeRow)
        
        #İÇERİĞİ TEMİZLE
        self.pushButton_8 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_8.setGeometry(QtCore.QRect(580, 60, 31, 31))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_8.setFont(font)
        self.pushButton_8.setObjectName("pushButton_8")
        self.pushButton_8.clicked.connect(self.tablotemizle)
        self.updateTable()
       
       
        #RAPOR TABLOSU
        self.tableWidget_2 = QtWidgets.QTableWidget(parent=self.centralwidget)
        self.tableWidget_2.setGeometry(QtCore.QRect(610, 0, 481, 501))
        self.tableWidget_2.setObjectName("tableWidget_2")
        self.tableWidget_2.setColumnCount(0)
        self.tableWidget_2.setRowCount(0)
        
        
        
        
        # HESAPLA BUTONU
        self.pushButton_23 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_23.setGeometry(QtCore.QRect(440, 340, 61, 31))
        self.pushButton_23.setObjectName("pushButton_23")
        self.pushButton_23.clicked.connect(self.hesaplamalar)

        
        #A OLUŞTUR BUTONU
        self.pushButton_2 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(20, 380, 81, 71))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.clicked.connect(self.katsayilarmatrisi)
    
        #L OLUŞTUR BUTONU
        self.pushButton_3 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_3.setGeometry(QtCore.QRect(20, 460, 81, 71))
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_3.clicked.connect(self.olcuvektoru)
        
        #P OLUŞTUR BUTONU
        self.pushButton = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(20, 540, 81, 71))
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.agirlikmatrisi)
        
        
        # ATP GÖRÜNTÜLE BUTONU
        self.pushButton_11 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_11.setGeometry(QtCore.QRect(130, 380, 81, 71))
        self.pushButton_11.setObjectName("pushButton_11")
        self.pushButton_11.clicked.connect(self.ATP)

        # N-1 GÖRÜNTÜLE BUTONU
        self.pushButton_9 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_9.setGeometry(QtCore.QRect(130, 460, 81, 71))
        self.pushButton_9.setObjectName("pushButton_9")
        self.pushButton_9.clicked.connect(self.N1)
        
        # vTPv GÖRÜNTÜLE BUTONU
        self.pushButton_10 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_10.setGeometry(QtCore.QRect(130, 540, 81, 71))
        self.pushButton_10.setObjectName("pushButton_10")
        self.pushButton_10.clicked.connect(self.vTPv)

        # N GÖRÜNTÜLE BUTONU
        self.pushButton_12 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_12.setGeometry(QtCore.QRect(240, 380, 81, 71))
        self.pushButton_12.setObjectName("pushButton_12")
        self.pushButton_12.clicked.connect(self.N)

        # x GÖRÜNTÜLE BUTONU
        self.pushButton_14 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_14.setGeometry(QtCore.QRect(240, 460, 81, 71))
        self.pushButton_14.setObjectName("pushButton_14")
        self.pushButton_14.clicked.connect(self.x)

        # m0 GÖRÜNTÜLE BUTONU
        self.pushButton_13 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_13.setGeometry(QtCore.QRect(240, 540, 81, 71))
        self.pushButton_13.setObjectName("pushButton_13")
        self.pushButton_13.clicked.connect(self.m0)

        # n GÖRÜNTÜLE BUTONU
        self.pushButton_17 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_17.setGeometry(QtCore.QRect(350, 380, 81, 71))
        self.pushButton_17.setObjectName("pushButton_17")
        self.pushButton_17.clicked.connect(self.n)

        # v GÖRÜNTÜLE BUTONU
        self.pushButton_15 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_15.setGeometry(QtCore.QRect(350, 460, 81, 71))
        self.pushButton_15.setObjectName("pushButton_15")
        self.pushButton_15.clicked.connect(self.v)

        # mX GÖRÜNTÜLE BUTONU
        self.pushButton_16 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_16.setGeometry(QtCore.QRect(350, 540, 81, 71))
        self.pushButton_16.setObjectName("pushButton_16")
        self.pushButton_16.clicked.connect(self.mX)

        # mY GÖRÜNTÜLE BUTONU
        self.pushButton_19 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_19.setGeometry(QtCore.QRect(460, 380, 81, 71))
        self.pushButton_19.setObjectName("pushButton_19")
        self.pushButton_19.clicked.connect(self.mY)

        # mZ GÖRÜNTÜLE BUTONU
        self.pushButton_20 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_20.setGeometry(QtCore.QRect(460, 460, 81, 71))
        self.pushButton_20.setObjectName("pushButton_20")
        self.pushButton_20.clicked.connect(self.mZ)

        #aTPv GÖRÜNTÜLE BUTONU
        self.pushButton_18 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_18.setGeometry(QtCore.QRect(460, 540, 81, 71))
        self.pushButton_18.setObjectName("pushButton_18")
        self.pushButton_18.clicked.connect(self.aTPv)

        #UPDATE GÖRSEL GİF
        self.labelresim_3 = QLabel(parent=self.centralwidget)
        self.labelresim_3.setGeometry(QtCore.QRect(10, 610, 221, 201))
        self.labelresim_3.setCursor(Qt.CursorShape.PointingHandCursor)
         

        # matrix GİF
        self.labelresim = QLabel(parent=self.centralwidget)
        self.labelresim.setGeometry(QtCore.QRect(580, 90, 31, 521))
        self.labelresim_3.setCursor(Qt.CursorShape.PointingHandCursor)
        image_path = r"images\matrix.gif"
        movie = QMovie(image_path)
        self.labelresim.setMovie(movie)
        self.labelresim.setScaledContents(True)
        movie.start()


        #RAPOR OLUŞTUR
        self.pushButton_21 = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton_21.setGeometry(QtCore.QRect(790, 510, 41, 31))
        self.pushButton_21.setObjectName("pushButton_21")
        self.pushButton_21.clicked.connect(self.load_excel)
        
        

        self.labelresim_2 = QLabel(parent=self.centralwidget)
        self.labelresim_2.setGeometry(QtCore.QRect(840, 500, 251, 111))
        self.labelresim_2.setCursor(Qt.CursorShape.PointingHandCursor)
        image_path_2 = r"images\exit.gif"
        movie_2 = QMovie(image_path_2)
        self.labelresim_2.setMovie(movie_2)
        self.labelresim_2.setScaledContents(True)
        self.labelresim_2.mousePressEvent = self.exit_application
        movie_2.start()
        
        
        
        
        self.label = QtWidgets.QLabel(parent=self.centralwidget)
        self.label.setGeometry(QtCore.QRect(10, 340, 101, 41))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(10)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(parent=self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(290, 340, 111, 41))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(10)
        font.setBold(True)
        font.setUnderline(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_4 = QtWidgets.QLabel(parent=self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(620, 500, 181, 51))
        font = QtGui.QFont()
        font.setFamily("Calibri Light")
        font.setPointSize(10)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")

        
        
        self.line_2 = QtWidgets.QFrame(parent=self.centralwidget)
        self.line_2.setGeometry(QtCore.QRect(100, 340, 41, 271))
        self.line_2.setFrameShape(QtWidgets.QFrame.Shape.VLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Shadow.Sunken)
        self.line_2.setObjectName("line_2")
       
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(parent=MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        self.setup_table_stylesheet()
        self.setup_table_stylesheet_2()
     

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.pushButton_4.setText(_translate("MainWindow", "+"))
        self.pushButton_5.setText(_translate("MainWindow", "+"))
        self.pushButton_6.setText(_translate("MainWindow", "-"))
        self.pushButton_7.setText(_translate("MainWindow", "-"))
        self.pushButton_8.setText(_translate("MainWindow", "SİL"))
        self.pushButton.setText(_translate("MainWindow", "P"))
        self.pushButton_2.setText(_translate("MainWindow", "A"))
        self.pushButton_3.setText(_translate("MainWindow", "L"))
        self.label.setText(_translate("MainWindow", "  OLUŞTUR  "))
        self.label_2.setText(_translate("MainWindow", "GÖRÜNTÜLE"))
        self.pushButton_9.setText(_translate("MainWindow", "N-1"))
        self.pushButton_10.setText(_translate("MainWindow", "vTPv"))
        self.pushButton_11.setText(_translate("MainWindow", "ATP"))
        self.pushButton_12.setText(_translate("MainWindow", "N"))
        self.pushButton_13.setText(_translate("MainWindow", "m0"))
        self.pushButton_14.setText(_translate("MainWindow", "x"))
        self.pushButton_15.setText(_translate("MainWindow", "v"))
        self.pushButton_16.setText(_translate("MainWindow", "mX"))
        self.pushButton_17.setText(_translate("MainWindow", "n"))
        self.pushButton_18.setText(_translate("MainWindow", "ATPv"))
        self.pushButton_19.setText(_translate("MainWindow", "mY"))
        self.pushButton_20.setText(_translate("MainWindow", "mZ"))
        self.label_4.setText(_translate("MainWindow", "Raporu Görüntüle -->"))
        self.pushButton_21.setText(_translate("MainWindow", "OK"))
        self.pushButton_23.setText(_translate("MainWindow", "HESAPLA"))
    
    
    def exit_application(self, event):
        QApplication.quit()
    
    
    def setup_table_stylesheet(self):
        table_style = (
            "QTableWidget { background-color: black; }"
            "QTableWidget QTableCornerButton::section { background-color: black; }"
            "QTableWidget::item { color: white; }"
            "QHeaderView::section { background-color: black; color: white; }"
        )

        self.tableWidget_2.setStyleSheet(table_style)
    
    def setup_table_stylesheet_2(self):
       
        table_style = (
            "QTableWidget { background-color: grey; }"
            "QTableWidget QTableCornerButton::section { background-color: grey; }"
            "QTableWidget::item { background-color: white ; color: black; }"
            "QHeaderView::section { background-color: brown; color: white; }"
        )
        self.tableWidget.setStyleSheet(table_style)    
    
   
   
    def check_matrices(self):
        if self.a_matris is None or self.l_matris is None or self.p_matris is None:
            QMessageBox.warning(self.centralwidget, "Uyarı", "Matrisler oluşturulmamış!")
            return False
        else:
                self.a_matris = np.array(self.a_matris)
                self.p_matris = np.array(self.p_matris)
                self.l_matris = np.array(self.l_matris)
            
             # A transpoze P matrisinin boyutlarını kontrol et
                try:
                    a = np.dot(self.a_matris.T, self.p_matris)
                    np.dot(a, self.l_matris)
                except ValueError:
                    QMessageBox.warning(self.centralwidget, "Uyarı", "ATPl matrisinin çarpımı yapıılamıyor! Matris boyutlarını kontrol edin.")
                    return False
                
                return True
    
   
   
    def get_table_data(self):
        matris = []
        for row in range(self.rowCount):
            row_data = []
            for column in range(self.columnCount):
                item = self.tableWidget.item(row, column)
                if item is not None and item.text():
                    row_data.append(float(item.text()))
                else:
                    row_data.append(0.0)
            matris.append(row_data)
        return matris
   
   
   
    def format_matrix(self, matris):
        formatted = ""
        for row in matris:
            formatted += "\t".join(map(str, row)) + "\n"
        return formatted
   
   
    def show_success_message(self, message):
        dialog = QMessageBox(self.centralwidget)
        dialog.setWindowTitle("Başarılı")
        dialog.setText(message)
        dialog.setIcon(QMessageBox.Icon.Information)
        dialog.exec()
   
   
   
    def katsayilarmatrisi(self):
        if self.a_matris is not None:
            confirm_dialog = QMessageBox(self.centralwidget)
            confirm_dialog.setWindowTitle("Mevcut Matrisi Değiştir")
            confirm_dialog.setText("Mevcut matrisi değiştirmek istediğinize emin misiniz?\n\nMevcut Matris:\n" + self.format_matrix(self.a_matris))
            confirm_dialog.setIcon(QMessageBox.Icon.Warning)
            yes_button = confirm_dialog.addButton("Evet", QMessageBox.ButtonRole.YesRole)
            no_button = confirm_dialog.addButton("Hayır", QMessageBox.ButtonRole.NoRole)
            confirm_dialog.exec()

            if confirm_dialog.clickedButton() == yes_button:
                self.a_matris = self.get_table_data()
                self.show_success_message("Matris değiştirildi!")
                self.a_matriskontrol = 1
        else:
            self.a_matris = self.get_table_data()
            self.show_success_message("Matris oluşturuldu!")
            self.a_matriskontrol = 1

    
    def agirlikmatrisi(self):
        if self.p_matris is not None:
            confirm_dialog = QMessageBox(self.centralwidget)
            confirm_dialog.setWindowTitle("Mevcut Ağırlık Matrisi Değiştir")
            confirm_dialog.setText("Mevcut ağırlık matrisini değiştirmek istediğinize emin misiniz?\n\nMevcut Matris:\n" + self.format_matrix(self.p_matris))
            confirm_dialog.setIcon(QMessageBox.Icon.Warning)
            yes_button = confirm_dialog.addButton("Evet", QMessageBox.ButtonRole.YesRole)
            no_button = confirm_dialog.addButton("Hayır", QMessageBox.ButtonRole.NoRole)
            confirm_dialog.exec()

            if confirm_dialog.clickedButton() == yes_button:
                self.p_matris = self.get_table_data()
                self.show_success_message("Ağırlık matrisi değiştirildi!")
                self.p_matriskontrol = 1
        else:
            self.p_matris = self.get_table_data()
            self.show_success_message("Ağırlık matrisi oluşturuldu!")
            self.p_matriskontrol = 1
    
    def olcuvektoru(self):
        if self.l_matris is not None:
            confirm_dialog = QMessageBox(self.centralwidget)
            confirm_dialog.setWindowTitle("Mevcut ölçü vektörünü Değiştir")
            confirm_dialog.setText("Mevcut ölçü vektörünü değiştirmek istediğinize emin misiniz?\n\nMevcut Matris:\n" + self.format_matrix(self.l_matris))
            confirm_dialog.setIcon(QMessageBox.Icon.Warning)
            yes_button = confirm_dialog.addButton("Evet", QMessageBox.ButtonRole.YesRole)
            no_button = confirm_dialog.addButton("Hayır", QMessageBox.ButtonRole.NoRole)
            confirm_dialog.exec()
            
            if confirm_dialog.clickedButton() == yes_button:
                self.l_matris = self.get_table_data()
                self.show_success_message("Ölçü Vektörü değiştirildi!")
                self.l_matriskontrol = 1
        else:
            self.l_matris = self.get_table_data()
            self.show_success_message("Ölçü Vektörü oluşturuldu!")
            self.l_matriskontrol = 1
   

   
   
   
    def tablotemizle(self):
        self.tableWidget.clearContents()
        self.updateTable()
        self.columnCount = 3
        self.rowCount = 3
        self.updateTable()
         
         
         
    def updateTable(self):
        self.tableWidget.setRowCount(self.rowCount)
        self.tableWidget.setColumnCount(self.columnCount)
         # Tablonun boyutunu güncelle
        


    def addColumn(self):
        self.columnCount += 1
        self.updateTable()

    def removeColumn(self):
        if self.columnCount > 1:
            self.columnCount -= 1
            self.updateTable()

    def addRow(self):
        self.rowCount += 1
        self.updateTable()

    def removeRow(self):
        if self.rowCount > 1:
            self.rowCount -= 1
            self.updateTable()



    
    def hesaplamalar(self):
        # Matrisleri kontrol et
        if not self.check_matrices():
            return

        

        self.a_matris = np.array(self.a_matris)
        self.p_matris = np.array(self.p_matris)
        self.l_matris = np.array(self.l_matris)
        
        
        try:
            self.a_matris_transpoze = self.a_matris.T
            
            # ATP
            self.ATP_matris = np.dot(self.a_matris_transpoze, self.p_matris)
            
            # N
            self.N_matris =  np.dot(self.ATP_matris, self.a_matris)   
        
            # n
            self.n_matris = np.dot(self.ATP_matris, self.l_matris)   
            
            # N-1
            self.N1_matris = np.linalg.inv(self.N_matris)

            # x
            self.x_matris = np.dot(self.N1_matris, self.n_matris) 
            self.bilinmeyensayisi = self.x_matris.shape[0]  
            # v
            self.v_matris = np.dot(self.a_matris, self.x_matris) - self.l_matris

            # v transpoze
            self.v_matris_transpoze = self.v_matris.T
            
            #vTP
            self.vTP_matris = np.dot(self.v_matris_transpoze,self.p_matris)
            
            #vTPv
            self.vTPv_matris = np.dot(self.vTP_matris, self.v_matris)
            
            #m0
            self.olcusayisi = self.a_matris.shape[0]
            self.m0_matris = math.sqrt((self.vTPv_matris[0, 0]) / (self.olcusayisi - self.bilinmeyensayisi))

            #mX
            self.mX_matris = self.m0_matris * math.sqrt( self.N1_matris[0,0])

            #mY
            self.mY_matris = self.m0_matris * math.sqrt( self.N1_matris[1,1] )

            
            if self.x_matris.shape[0] > 2:
                #mZ
                self.mZ_matris = self.m0_matris * math.sqrt( self.N1_matris[2,2])

            #ATPv
            self.ATPv_matris = np.dot(self.ATP_matris, self.v_matris)
            self.excel_kaydet()
            self.show_success_message("Hesaplama Başarılı. Rapor Oluşturuldu")
        except np.linalg.LinAlgError:
            QMessageBox.warning(self.centralwidget, "Hata", "Matris çarpımı veya tersi alınamıyor!")
        
        
        
    def ATP(self):
        if self.ATP_matris is not None:
            self.tableWidget.setRowCount(len(self.ATP_matris))
            self.tableWidget.setColumnCount(len(self.ATP_matris[0]))
            
            for row in range(len(self.ATP_matris)):
                for column in range(len(self.ATP_matris[0])):
                    item = QtWidgets.QTableWidgetItem(str(self.ATP_matris[row][column]))
                    self.tableWidget.setItem(row, column, item)
        else:
            QMessageBox.warning(self.centralwidget, "Uyarı", "ATP Matrisi henüz oluşturulmadı.")
    
    def N(self):
         if self.N_matris is not None:
            self.tableWidget.setRowCount(len(self.N_matris))
            self.tableWidget.setColumnCount(len(self.N_matris[0]))
            
            for row in range(len(self.N_matris)):
                for column in range(len(self.N_matris[0])):
                    item = QtWidgets.QTableWidgetItem(str(self.N_matris[row][column]))
                    self.tableWidget.setItem(row, column, item)
         else:
            QMessageBox.warning(self.centralwidget, "Uyarı", "N Matrisi henüz oluşturulmadı.")
    
    def N1(self):
        if self.N1_matris is not None:
            self.tableWidget.setRowCount(len(self.N1_matris))
            self.tableWidget.setColumnCount(len(self.N1_matris[0]))
            
            for row in range(len(self.N1_matris)):
                for column in range(len(self.N1_matris[0])):
                    item = QtWidgets.QTableWidgetItem(str(self.N1_matris[row][column]))
                    self.tableWidget.setItem(row, column, item)
        else:
            QMessageBox.warning(self.centralwidget, "Uyarı", "N-1 Matrisi henüz oluşturulmadı.")
    
    def n(self):
        if self.n_matris is not None:
            self.tableWidget.setRowCount(len(self.n_matris))
            self.tableWidget.setColumnCount(len(self.n_matris[0]))
            
            for row in range(len(self.n_matris)):
                for column in range(len(self.n_matris[0])):
                    item = QtWidgets.QTableWidgetItem(str(self.n_matris[row][column]))
                    self.tableWidget.setItem(row, column, item)
        else:
            QMessageBox.warning(self.centralwidget, "Uyarı", "n Matrisi henüz oluşturulmadı.")
    
    def x(self):
        if self.x_matris is not None:
            self.tableWidget.setRowCount(len(self.x_matris))
            self.tableWidget.setColumnCount(len(self.x_matris[0]))
            
            for row in range(len(self.x_matris)):
                for column in range(len(self.x_matris[0])):
                    item = QtWidgets.QTableWidgetItem(str(self.x_matris[row][column]))
                    self.tableWidget.setItem(row, column, item)
                    
                           
                    
        else:
            QMessageBox.warning(self.centralwidget, "Uyarı", "x Matrisi henüz oluşturulmadı.")
    
    def vTPv(self):
        if self.vTPv_matris is not None:
            self.tableWidget.setRowCount(len(self.vTPv_matris))
            self.tableWidget.setColumnCount(len(self.vTPv_matris[0]))
            
            for row in range(len(self.vTPv_matris)):
                for column in range(len(self.vTPv_matris[0])):
                    item = QtWidgets.QTableWidgetItem(str(self.vTPv_matris[row][column]))
                    self.tableWidget.setItem(row, column, item)
        else:
            QMessageBox.warning(self.centralwidget, "Uyarı", "vTPv Matrisi henüz oluşturulmadı.")
    
    def v(self):
        if self.v_matris is not None:
            self.tableWidget.setRowCount(len(self.v_matris))
            self.tableWidget.setColumnCount(len(self.v_matris[0]))
            
            for row in range(len(self.v_matris)):
                for column in range(len(self.v_matris[0])):
                    item = QtWidgets.QTableWidgetItem(str(self.v_matris[row][column]))
                    self.tableWidget.setItem(row, column, item)
        else:
            QMessageBox.warning(self.centralwidget, "Uyarı", "v Matrisi henüz oluşturulmadı.")
    
    def m0(self):
         
        if self.m0_matris is not None:
            self.tableWidget.setRowCount(1)
            self.tableWidget.setColumnCount(1)
            
            item = QtWidgets.QTableWidgetItem(str(self.m0_matris))
            self.tableWidget.setItem(0, 0, item)
        else:
            QMessageBox.warning(self.centralwidget, "Uyarı", "m0 Matrisi henüz oluşturulmadı.")

    def mX(self):
        if self.mX_matris is not None:
            self.tableWidget.setRowCount(1)
            self.tableWidget.setColumnCount(1)
            
            item = QtWidgets.QTableWidgetItem(str(self.mX_matris))
            self.tableWidget.setItem(0, 0, item)
        else:
            QMessageBox.warning(self.centralwidget, "Uyarı", "mX Matrisi henüz oluşturulmadı.")

    def mY(self):
        if self.mY_matris is not None:
            self.tableWidget.setRowCount(1)
            self.tableWidget.setColumnCount(1)
            
            item = QtWidgets.QTableWidgetItem(str(self.mY_matris))
            self.tableWidget.setItem(0, 0, item)
        else:
            QMessageBox.warning(self.centralwidget, "Uyarı", "mY Matrisi henüz oluşturulmadı.")

    def mZ(self):
        if self.mZ_matris is not None:
            self.tableWidget.setRowCount(1)
            self.tableWidget.setColumnCount(1)
            
            item = QtWidgets.QTableWidgetItem(str(self.mZ_matris))
            self.tableWidget.setItem(0, 0, item)
        else:
            QMessageBox.warning(self.centralwidget, "Uyarı", "mZ Matrisi henüz oluşturulmadı.")

    def aTPv(self):
        if self.ATPv_matris is not None:
            self.tableWidget.setRowCount(len(self.ATPv_matris))
            self.tableWidget.setColumnCount(len(self.ATPv_matris[0]))
            
            for row in range(len(self.ATPv_matris)):
                for column in range(len(self.ATPv_matris[0])):
                    item = QtWidgets.QTableWidgetItem(str(self.ATPv_matris[row][column]))
                    self.tableWidget.setItem(row, column, item)
        else:
            QMessageBox.warning(self.centralwidget, "Uyarı", "ATPv Matrisi henüz oluşturulmadı.")   


    def excel_kaydet(self):
       
        self.ATP_matris = np.array(self.ATP_matris)
        self.N_matris = np.array(self.N_matris)
        self.n_matris  = np.array(self.n_matris)
        self.N1_matris = np.array(self.N1_matris)
        self.x_matris = np.array(self.x_matris)
        self.v_matris = np.array(self.v_matris)
        self.vTPv_matris = np.array(self.vTPv_matris)
        self.m0_matris = [str(self.m0_matris)]
        self.mX_matris = [str(self.mX_matris)]
        self.mY_matris = [str(self.mY_matris)]
        self.mZ_matris = [str(self.mZ_matris)]
        self.ATPv_matris = np.array(self.ATPv_matris)

        
        data = [
                (' ', ' ', ' '),
                (' ', ' ', ' '),
                ('ATP', 'Değer', self.ATP_matris),
                ('N', 'Değer', self.N_matris),
                ('n', 'Değer', self.n_matris),
                ('N-1', 'Değer', self.N1_matris),
                ('x', 'Değer', self.x_matris),
                ('v', 'Değer', self.v_matris),
                ('vTPv', 'Değer', self.vTPv_matris),
                ('ATPv', 'Değer', self.ATPv_matris),
    
           
        ]


        row_number = 10
        dosya_adı = 'Dengeleme Raporu.xlsx'
        if os.path.isfile(dosya_adı):
            os.remove(dosya_adı)
        dosya_mevcut = os.path.isfile(dosya_adı)

        if not dosya_mevcut:
            workbook = openpyxl.Workbook()
        else:
            workbook = openpyxl.load_workbook(dosya_adı)

        sheet = workbook.active
        for variable, description, matrix in data:
            # Append variable name and description
            sheet.append([variable, description, None])
            row_number += 1
            # Insert the matrix elements into separate cells
            for matris_satir in matrix:
                row = []
                for eleman in matris_satir:
                    row.append(eleman)
                sheet.append(row)
                row_number += 1
            # Add a blank row between matrices
            sheet.append([])

        for col_num, value in enumerate(['Değişken', 'Açıklama', 'Değer']):
            sheet.cell(row=1, column=col_num+1, value=value)
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = length
        
        
        sutun = 'A'
        satir = row_number
        hucre_adresi = f'{sutun}{satir}'
        sheet[hucre_adresi] = 'm0 = '
        sutun = 'A'
        satir = row_number + 1
        hucre_adresi = f'{sutun}{satir}'
        sheet[hucre_adresi] = 'mX = '
        sutun = 'A'
        satir = row_number + 2
        hucre_adresi = f'{sutun}{satir}'
        sheet[hucre_adresi] = 'mY = ' 
        sutun = 'A'
        satir = row_number + 3
        hucre_adresi = f'{sutun}{satir}'
        sheet[hucre_adresi] = 'mZ = '
        
        sutun = 'B'
        satir = row_number
        hucre_adresi = f'{sutun}{satir}'
        sheet[hucre_adresi] = self.m0_matris[0]
        sutun = 'B'
        satir = row_number + 1
        hucre_adresi = f'{sutun}{satir}'
        sheet[hucre_adresi] = self.mX_matris[0]
        sutun = 'B'
        satir = row_number + 2
        hucre_adresi = f'{sutun}{satir}'
        sheet[hucre_adresi] = self.mY_matris[0]
        sutun = 'B'
        satir = row_number + 3
        hucre_adresi = f'{sutun}{satir}'
        sheet[hucre_adresi] = self.mZ_matris[0]    
       
        workbook.save(dosya_adı)


    def load_excel(self):
        print(1)
        file_name, _ = QtWidgets.QFileDialog.getOpenFileName(self.centralwidget, "Excel Dosyasını Seç", "", "Excel Dosyaları (*.xlsx *.xls);;Tüm Dosyalar (*)")
        if file_name:
            import openpyxl
            self.excel_file_path = file_name  # Excel dosyasının yolunu sakla
            # Excel dosyasını aç
            workbook = openpyxl.load_workbook(file_name)
            sheet = workbook.active
            # Tabloyu temizle ve başlıkları ekleyin
            self.tableWidget_2.clear()
            self.tableWidget_2.setRowCount(0)  # Tabloyu sıfırla

            header_labels = [cell.value for cell in sheet[1]]  # İlk satır başlıklar
            self.tableWidget_2.setColumnCount(len(header_labels))
            self.tableWidget_2.setHorizontalHeaderLabels(header_labels)
            # Excel dosyasındaki verileri tabloya aktar
            for row in sheet.iter_rows(min_row=2):  # İlk satır başlıklar olduğu için 2'den başlatıyoruz
                row_data = [cell.value if cell.value is not None else '' for cell in row]
                row_position = self.tableWidget_2.rowCount()  # Mevcut satır sayısı
                self.tableWidget_2.insertRow(row_position)
                for col_position, cell_value in enumerate(row_data):
                    item = QtWidgets.QTableWidgetItem(str(cell_value))
                    self.tableWidget_2.setItem(row_position, col_position, item)

            workbook.close()
            self.tableWidget_2.update()
       


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec())
